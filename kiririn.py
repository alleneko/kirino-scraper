import time

import openpyxl

from solaris import Solaris

def get_last_row(worksheet, column):
	selected_column = worksheet[column]
	for row in range(1, worksheet.max_row):
		if selected_column[row].value == None:
			return row

spreadsheet_path = "./anime_figures.xlsx"

workbook = openpyxl.load_workbook(spreadsheet_path)
workbook.active = workbook["SolarisJapan"]

sheet = workbook.active
last_figure = sheet.max_row

solaris = Solaris()
for figure in range(1, last_figure - 1):
	url = sheet["B"][figure].value
	solaris.get_page(url)
	time.sleep(10)
	solaris.get_buttons()
	sheet["C"][figure].value = solaris.prices["new"]
	sheet["D"][figure].value = solaris.prices["used"]
	sheet["E"][figure].value = solaris.prices["pre_order"]
	solaris.get_shipping()
	sheet["F"][figure].value = solaris.shipping
	workbook.save(spreadsheet_path)
	time.sleep(30)